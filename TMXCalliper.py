# Modules

from appJar import gui
from pathlib import Path
import openpyxl
import xml.etree.ElementTree as ET
import Levenshtein
import datetime
import os
from openpyxl.styles import PatternFill, Font, Fill, Alignment, Border, Side

# Variables
version = 0.6
src_my_file = None
dst_my_file = None
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
colors = ['5AA0C0', '6CB8C9', '7ECED2', '91DAD4', 'A4E2D4', 'B7E9D8']

# Parsing TMX


def parseTMX(my_fileName):
    tree = ET.parse(my_fileName)
    root = tree.getroot()
    bilingual_table_from_xml = []
    my_dictionary = []
    x = 0
    for segments in root.iter('tuv'):
        singleSegment = segments.find('seg').text
        if singleSegment:
            bilingual_table_from_xml.append(singleSegment)
    my_dictionary = {}
    for source, target in zip(
            bilingual_table_from_xml[0::2], bilingual_table_from_xml[1::2]):
        my_dictionary[source] = target
    return(my_dictionary)

# Main proc


def main_proc(my_file1, my_file2):
    my_table1 = parseTMX(my_file1)
    my_table2 = parseTMX(my_file2)


# Create XLSX file

    wb = openpyxl.Workbook()
    sheet = wb['Sheet']

    sheet.cell(row=1, column=1).value = 'Source'
    sheet.cell(row=1, column=2).value = 'Target 1'
    sheet.cell(row=1, column=3).value = 'Target 2'
    sheet.cell(row=1, column=4).value = 'Distance'

    x = 2
    for key, value in my_table1.items():
        if key in my_table2:  # Compare only when the same segment occures in the 2nd TMX file
            sheet.cell(row=x, column=1).value = key
            sheet.cell(
                row=x,
                column=1).alignment = Alignment(
                wrapText=True,
                horizontal='left',
                vertical='center')
            sheet.cell(
                row=x, column=1).fill = PatternFill(
                "solid", fgColor='DFC45F')
            sheet.cell(row=x, column=1).border = thin_border
            sheet.cell(row=x, column=2).value = value
            sheet.cell(
                row=x,
                column=2).alignment = Alignment(
                wrapText=True,
                horizontal='left',
                vertical='center')
            sheet.cell(row=x, column=2).border = thin_border
            second_version = my_table2[key]
            sheet.cell(row=x, column=3).value = second_version
            sheet.cell(
                row=x,
                column=3).alignment = Alignment(
                wrapText=True,
                horizontal='left',
                vertical='center')
            sheet.cell(row=x, column=3).border = thin_border
            editdistance = Levenshtein.distance(value, second_version)
            average_length = (len(value) + len(second_version) / 2)
            div_length = (editdistance / average_length)
            sheet.cell(row=x, column=4).value = div_length
            sheet.cell(
                row=x,
                column=4).alignment = Alignment(
                wrapText=True,
                horizontal='center',
                vertical='center')
            sheet.cell(row=x, column=4).border = thin_border
            sheet.cell(row=x, column=4).number_format = '0.00'
            color_number = (4 - int((editdistance / average_length) / 0.2))
            if div_length != 0:
                sheet.cell(
                    row=x, column=2).fill = PatternFill(
                    "solid", fgColor=colors[color_number])
                sheet.cell(
                    row=x, column=3).fill = PatternFill(
                    "solid", fgColor=colors[color_number])
                sheet.cell(
                    row=x, column=4).fill = PatternFill(
                    "solid", fgColor=colors[color_number])
            x = x + 1
    textsredniej = '= AVERAGE(D2:D' + str(x - 1) + ')'
    sheet.cell(row=x, column=4).value = textsredniej
    sheet.cell(row=x, column=4).number_format = '0.00'
    sheet.cell(
        row=x,
        column=4).alignment = Alignment(
        wrapText=True,
        horizontal='center',
        vertical='center')
    sheet.cell(
        row=x,
        column=3).alignment = Alignment(
        wrapText=True,
        horizontal='right',
        vertical='center')
    sheet.cell(row=x, column=3).value = 'Average dist.'
    now = datetime.datetime.now()
    timestamp = (now.strftime("%Y%m%d-%H%M%S"))
    savename = timestamp + '_TMXCalliper_v.' + str(version) + '.xlsx'
    wb.save(savename)

# GUI


def press(button):
    if button == "Process":
        src_my_file = os.path.normpath(app.getEntry("Input_my_file1"))
        dst_my_file = os.path.normpath(app.getEntry("Input_my_file2"))
        main_proc(src_my_file, dst_my_file)
        app.infoBox(
            "Info",
            "Files have been succesfully processed!",
            parent=None)
        app.stop()
    else:
        app.stop()


app = gui("TMXCalliper", useTtk=True)
app.setTtkTheme("default")
app.setSize(600, 150)
app.addLabel(" Choose TMX file #1:")
app.addFileEntry("Input_my_file1")
app.addLabel(" Choose TMX file #2:")
app.addFileEntry("Input_my_file2")
app.addButtons(["Process", "Quit"], press)
app.go()
